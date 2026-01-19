import openpyxl
import re

def get_techniques_xlsx(path, sheet_name=None):
    wb = openpyxl.load_workbook(path)
    
    # Si no se proporciona el nombre de la hoja, se usa la activa
    ws = wb[sheet_name] if sheet_name else wb.active

    diccionario = {}

    for fila in ws.iter_rows(min_row=2, values_only=True):  # min_row=2 para omitir encabezado
        valor_col1 = fila[0]
        clave_col2 = fila[1]
        valor_col3 = 1 if fila[2] == 'clasificador' else 0

        if clave_col2 is not None:
            diccionario[clave_col2] = (valor_col1, valor_col3)

    return diccionario

def get_datasets(path, sheet_name=None):
    wb = openpyxl.load_workbook(path)
    
    # Si no se proporciona el nombre de la hoja, se usa la activa
    ws = wb[sheet_name] if sheet_name else wb.active

    diccionario = {}

    for fila in ws.iter_rows(min_row=2, values_only=True):  # min_row=2 para omitir encabezado
        valor_col1 = fila[0]
        valor_col3 = fila[2]
        valor_col4 = fila[3]
        valor_col5 = fila[4]
        valor_col6 = fila[5]
        none_value = valor_col1 is None or valor_col3 is None or valor_col4 is None or valor_col5 is None or valor_col6 is None or valor_col3 != 2

        if not none_value and type(valor_col4) is str and ':' in valor_col4:
            texto_limpio = re.sub(r'\([^)]*\)', '', valor_col4)
            texto_limpio = texto_limpio.split(': ')[1]
            texto_limpio = texto_limpio.strip()
            secciones = texto_limpio.split(', ')
            discretos = continuos = categoricos = 0
            for seccion in secciones:
                if ' D' in seccion:
                    discretos = int(seccion.split(' D')[0].strip())
                elif ' CO' in seccion:
                    continuos = int(seccion.split(' CO')[0].strip())
                elif ' CA' in seccion:
                    categoricos = int(seccion.split(' CA')[0].strip())

            diccionario[valor_col1] = (valor_col3, categoricos, continuos, discretos, valor_col5, valor_col6)

    return diccionario