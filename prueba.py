import json
import tkinter as tk
from tkinter import messagebox, ttk
import openpyxl
from datetime import datetime
import os
import re
import csv
import random

model_set = {
    "AdaBoost", "ADASYN", "ANN", "ASUWO", "BalanceCascade", "Borderline2SMOTE",
    "BPNN", "BSMOTE", "BWELM", "CCR-ELM", "ClusterSMOTE", "CS-ELM", "CSKELM",
    "CSMOTE", "DEBOHID", "D-ENN", "DNN", "DT", "EasyEnsemble", "EFSVM", "ELM",
    "HABC-WELM", "IWS-SMOTE", "KELM", "KMSMOTE", "kNN", "KNNOR", "KWELM", "LDA",
    "MLP", "MSOSS", "MWMOTE", "NB", "NGSMOTE", "Original", "RF", "ROS",
    "RUSBoost", "Safelevel", "SGO", "SMOTE", "SMOTE-ENN", "SMOTE-IPF", "S-RSB",
    "SSMOTE", "S-TL", "SVM", "SyM", "UBKELM-MV", "UBKELM-SV", "WELM", "WKSMOTE",
    "XGB", "BNB", "UFIDSF", "UCF", "RENN", "ENN", "AKNN", "GNB", "PAU", "OSS",
    "TL", "RBU", "RSDS", "SDUS"
}

# Reuse parsing classes
class Result:
    def __init__(self, value):
        if value == "" or value is None:
            self.value = [None]
            self.technique = [None]
            return
        values = value.split(" (")
        if len(values) > 1:
            number = float(values[0].replace(",", "."))
            self.value = [number if number <= 1 else number / 100]
            self.technique = [values[1].replace(")", "")]
        else:
            number = float(value.replace(",", "."))
            self.value = [number if number <= 1 else number / 100]
            self.technique = [None]

    def add(self, value):
        if value == "" or value is None:
            self.value.append(None)
            self.technique.append(None)
            return
        values = value.split(" (")
        if len(values) > 1:
            number = float(values[0].replace(",", "."))
            self.value.append(number if number <= 1 else number / 100)
            self.technique.append(values[1].replace(")", ""))
        else:
            number = float(value.replace(",", "."))
            self.value.append(number if number <= 1 else number / 100)
            self.technique.append(None)

class Dataset:
    def __init__(self, accuracy, f1, precision, recall, gmean, auc):
        self.accuracy = Result(accuracy)
        self.F1 = Result(f1)
        self.precision = Result(precision)
        self.recall = Result(recall)
        self.gmean = Result(gmean)
        self.auc = Result(auc)

    def add(self, accuracy, f1, precision, recall, gmean, auc):
        self.accuracy.add(accuracy)
        self.F1.add(f1)
        self.precision.add(precision)
        self.recall.add(recall)
        self.gmean.add(gmean)
        self.auc.add(auc)

class App:
    def __init__(self, root):
        root.title("Article to JSON Converter")
        self.datasets = {}

        # Article Frame
        art_frame = ttk.LabelFrame(root, text="Article Information", padding=10)
        art_frame.grid(row=0, column=0, padx=10, pady=5, sticky="ew")

        labels = ["Article ID", "DOI", "Title", "Date", "Keywords", "Impact Factor", "Citation Index"]
        self.entries = {}
        for i, lbl in enumerate(labels):
            ttk.Label(art_frame, text=lbl + ":").grid(row=i, column=0, sticky="e")
            ent = ttk.Entry(art_frame, width=40)
            ent.grid(row=i, column=1, pady=2, sticky="w")
            self.entries[lbl.lower().replace(" ", "_")] = ent

        # Dataset Frame
        ds_frame = ttk.LabelFrame(root, text="Datasets", padding=10)
        ds_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")

        ds_labels = ["Dataset ID", "Accuracy", "F1", "Precision", "Recall", "Gmean", "AUC"]
        self.ds_entries = {}
        for i, lbl in enumerate(ds_labels):
            ttk.Label(ds_frame, text=lbl + ":").grid(row=i, column=0, sticky="e")
            ent = ttk.Entry(ds_frame, width=30)
            ent.grid(row=i, column=1, pady=2, sticky="w")
            self.ds_entries[lbl.lower()] = ent

        self.add_btn = ttk.Button(ds_frame, text="Add Dataset", command=self.add_dataset)
        self.add_btn.grid(row=len(ds_labels), column=0, columnspan=2, pady=5)

        # Listbox to show added datasets
        ttk.Label(ds_frame, text="Added Datasets:").grid(row=0, column=2, padx=(10,0), sticky="nw")
        self.ds_list = tk.Listbox(ds_frame, height=8)
        self.ds_list.grid(row=1, column=2, rowspan=len(ds_labels), padx=(10,0), sticky="n")

        # Save Button
        self.save_btn = ttk.Button(root, text="Save Article to JSON", command=self.save_article)
        self.save_btn.grid(row=2, column=0, pady=10)

    def add_dataset(self):
        try:
            ds_id = self.ds_entries['dataset id'].get().strip()
            if not ds_id:
                raise ValueError("Dataset ID required")
            values = {k: v.get().strip() for k, v in self.ds_entries.items() if k != 'dataset id'}
            if int(ds_id) in self.datasets:
               self.datasets[int(ds_id)].add(**values)
            else:
                self.datasets[int(ds_id)] = Dataset(**values)
            self.ds_list.insert(tk.END, f"ID {ds_id}")
            # Clear dataset entries
            for ent in self.ds_entries.values():
                ent.delete(0, tk.END)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def save_article(self):
        try:
            # Gather article info
            info = {k: v.get().strip() for k, v in self.entries.items()}
            article_id = info['article_id']
            # Build Article-like dict
            article = {
                'doi': info['doi'],
                'tittle': info['title'],
                'date': info['date'],
                'keywords': info['keywords'].split(', '),
                'IF': float(info['impact_factor'].replace(',', '.')),  
                'CI': float(info['citation_index'].replace(',', '.')),  
                'datasets': {
                    k: {
                        'accuracy': v.accuracy.__dict__,
                        'F1': v.F1.__dict__,
                        'precision': v.precision.__dict__,
                        'recall': v.recall.__dict__,
                        'gmean': v.gmean.__dict__,
                        'auc': v.auc.__dict__
                    } for k, v in self.datasets.items()
                }
            }
            # Save to file
            with open(f"{article_id}.json", 'w', encoding='utf-8') as f:
                json.dump(article, f, indent=4, ensure_ascii=False)
            messagebox.showinfo("Success", f"Article saved to {article_id}.json")
            self.reset_all()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def reset_all(self):
        # clear article entries
        for ent in self.entries.values():
            ent.delete(0, tk.END)
        # clear datasets
        self.datasets.clear()
        self.ds_list.delete(0, tk.END)


class Auto:
    def __init__(self):
        self.datasets = {}
        self.inicio = True
        self.article = None
        self.doi = None
        self.tittle = None
        self.date = None
        self.keywords = None
        self.impact_factor = None
        self.citation_index = None

    def add_dataset(self, line):
        try:
            if line[0] is not None:
                if not self.inicio:
                    self.save_article()
                self.inicio = False
                self.datasets = {}
                self.article = line[0].strip()
                self.doi = line[1].strip()
                self.tittle = line[2].strip()
                self.date = line[3]
                self.keywords = line[4].strip()
                self.impact_factor = line[5] if line[5] is not None else "0.0"
                self.citation_index = line[6] if line[6] is not None else "0.0"
            ds_id = line[7]
            if not ds_id:
                raise ValueError("Dataset ID required")
            if int(ds_id) in self.datasets:
               self.datasets[int(ds_id)].add(line[8], line[9], line[10], line[11], line[12], line[13])
            else:
                self.datasets[int(ds_id)] = Dataset(line[8], line[9], line[10], line[11], line[12], line[13])
        except Exception as e:
            print(f"Error adding dataset: {e}")

    def save_article(self):
        try:
            # Gather article info
            article_id = self.article
            # Build Article-like dict
            article = {
                'doi': self.doi,
                'tittle': self.tittle,
                'date': self.date.strftime("%m/%Y"),
                'keywords': self.keywords.split(', '),
                'IF': float(self.impact_factor),  
                'CI': float(self.citation_index),  
                'datasets': {
                    k: {
                        'accuracy': v.accuracy.__dict__,
                        'F1': v.F1.__dict__,
                        'precision': v.precision.__dict__,
                        'recall': v.recall.__dict__,
                        'gmean': v.gmean.__dict__,
                        'auc': v.auc.__dict__
                    } for k, v in self.datasets.items()
                }
            }
            # Save to file
            with open(f"{article_id}.json", 'w', encoding='utf-8') as f:
                json.dump(article, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving article: {e}")

    def reset_all(self):
        # clear article entries
        for ent in self.entries.values():
            ent.delete(0, tk.END)
        # clear datasets
        self.datasets.clear()
        self.ds_list.delete(0, tk.END)


def leer_xlsx_y_contar(ruta_archivo):
    try:
        wb = openpyxl.load_workbook(ruta_archivo)
        hoja = wb.active  # Usa la hoja activa por defecto
        auto = Auto()

        for fila in hoja.iter_rows(min_row=1, max_col=14, values_only=True):
            auto.add_dataset(fila)
        
        auto.save_article()
    except Exception as e:
        print(f"Error al leer el archivo: {e}")

def cargar_jsons_desde_directorio(directorio="JSON"):
    jsons_cargados = []

    if not os.path.isdir(directorio):
        print(f"El directorio '{directorio}' no existe.")
        return jsons_cargados

    for nombre_archivo in os.listdir(directorio):
        if nombre_archivo.endswith(".json"):
            ruta_completa = os.path.join(directorio, nombre_archivo)
            try:
                with open(ruta_completa, 'r', encoding='utf-8') as archivo:
                    contenido = json.load(archivo)
                    contenido['article_id'] = nombre_archivo[:-5]  # Agregar ID del artículo
                    jsons_cargados.append(contenido)
            except Exception as e:
                print(f"Error al cargar {nombre_archivo}: {e}")

    print(f"{len(jsons_cargados)} archivos JSON cargados desde '{directorio}'.")
    return jsons_cargados

def get_techniques(jsons, by_article=False):
    if not by_article:
        techniques = set()
        for json_data in jsons:
            datasets = json_data.get('datasets', {})
            for dataset in datasets.values():
                for key in ['accuracy', 'F1', 'precision', 'recall', 'gmean', 'auc']:
                    if key in dataset:
                        for technique in dataset[key].get('technique', []):
                            if technique:
                                techniques.add(technique)
        return sorted(techniques)
    else:
        techniques_by_article = {}
        for json_data in jsons:
            article_id = json_data.get('article_id', 'unknown')
            techniques = set()
            datasets = json_data.get('datasets', {})
            for dataset in datasets.values():
                for key in ['accuracy', 'F1', 'precision', 'recall', 'gmean', 'auc']:
                    if key in dataset:
                        for technique in dataset[key].get('technique', []):
                            if technique:
                                techniques.add(technique)
            techniques_by_article[article_id] = sorted(techniques)
        return techniques_by_article

def analizar_jsons(jsons, tecnicas_validas=None):
    if tecnicas_validas is None:
        print("No valid techniques provided for analysis.")
        return
    if not jsons:
        print("No JSON data to analyze.")
        return

    for json_data in jsons:
        #print(f"DOI: {json_data.get('doi', 'N/A')}")
        #print(f"Title: {json_data.get('tittle', 'N/A')}")
        #print(f"Date: {json_data.get('date', 'N/A')}")
        #print(f"Keywords: {', '.join(json_data.get('keywords', []))}")
        #print(f"Impact Factor: {json_data.get('IF', 'N/A')}")
        #print(f"Citation Index: {json_data.get('CI', 'N/A')}")
        invalid_techniques = set()
        datasets = json_data.get('datasets', {})
        for ds_id, dataset in datasets.items():
            #print(f"\tDataset ID: {ds_id}")
            for metric, result in dataset.items():
                techniques = result.get('technique', [])
                for technique in techniques:
                    if technique is not None:
                        sections = technique.split(';')
                        for section in sections:
                            section = section.split(',')
                            for sec in section:
                                if sec not in tecnicas_validas:
                                    invalid_techniques.add(sec)
        if invalid_techniques:
            print(f"Article ID: {json_data.get('article_id', 'unknown')}")
            print(f"\tInvalid techniques found: {', '.join(invalid_techniques)}")

def get_techniques_xlsx(path, sheet_name=None):
    wb = openpyxl.load_workbook(path)
    
    # Si no se proporciona el nombre de la hoja, se usa la activa
    ws = wb[sheet_name] if sheet_name else wb.active

    diccionario = {}

    for fila in ws.iter_rows(min_row=2, values_only=True):  # min_row=2 para omitir encabezado
        valor_col1 = fila[0]
        clave_col2 = fila[1]
        valor_col3 = 1 if fila[2] == 'clasificador' else 0

        if clave_col2 is not None:
            diccionario[clave_col2] = (valor_col1, valor_col3)

    return diccionario

def get_datasets(path, sheet_name=None):
    wb = openpyxl.load_workbook(path)
    
    # Si no se proporciona el nombre de la hoja, se usa la activa
    ws = wb[sheet_name] if sheet_name else wb.active

    diccionario = {}

    for fila in ws.iter_rows(min_row=2, values_only=True):  # min_row=2 para omitir encabezado
        valor_col1 = fila[0]
        valor_col3 = fila[2]
        valor_col4 = fila[3]
        valor_col5 = fila[4]
        valor_col6 = fila[5]
        none_value = valor_col1 is None or valor_col3 is None or valor_col4 is None or valor_col5 is None or valor_col6 is None

        if not none_value and type(valor_col4) is str and ':' in valor_col4:
            texto_limpio = re.sub(r'\([^)]*\)', '', valor_col4)
            texto_limpio = texto_limpio.split(': ')[1]
            texto_limpio = texto_limpio.strip()
            secciones = texto_limpio.split(', ')
            discretos = continuos = categoricos = 0
            for seccion in secciones:
                if ' D' in seccion:
                    discretos = int(seccion.split(' D')[0].strip())
                elif ' CO' in seccion:
                    continuos = int(seccion.split(' CO')[0].strip())
                elif ' CA' in seccion:
                    categoricos = int(seccion.split(' CA')[0].strip())

            diccionario[valor_col1] = (valor_col3, categoricos, continuos, discretos, valor_col5, valor_col6)

    return diccionario
                         
if __name__ == '__main__':
    mode = input("Menu:" \
                 "\n\t1. Datasets' registration app" \
                 "\n\t2. Datasets' auto detection" \
                 "\n\t3. Analyze JSON directory" \
                 "\n\t4. Get metrics' resume" \
                 "\n\t5. Analyze JSONs" \
                 "\n\t6. Get techniques from xlsx" \
                 "\nEnter mode: ").strip()
    if mode not in ['1', '2', '3', '4', '5', '6']:
        print("Invalid mode selected. Exiting.")
        exit()
    if mode == '1':
        root = tk.Tk()
        app = App(root)
        root.mainloop()
    elif mode == '2':
        leer_xlsx_y_contar('articulos.xlsx')
    elif mode == '3':
        json = cargar_jsons_desde_directorio('JSON')
        techniques = get_techniques(json, by_article=True)
        for article_id, techniques in techniques.items():
            print(f"Article ID: {article_id}, Techniques: {', '.join(techniques)}")
    elif mode == '4':
        json = cargar_jsons_desde_directorio('JSON')
        results = { 'accuracy': 0, 'F1': 0, 'precision': 0, 'recall': 0, 'gmean': 0, 'auc': 0 }
        for article in json:
            datasets = article.get('datasets', {})
            for dataset in datasets.values():
                for key in results.keys():
                    for value in dataset[key]['value']:
                        if value is not None:
                            results[key] += 1
        print("Metrics summary:")
        for key, count in results.items():
            print(f"{key.capitalize()}: {count} values found")
    elif mode == '5':
        tecnicas_validas = {'AdaBoost', 'ADASYN', 'ANN', 'ASUWO', 'BalanceCascade', 'Borderline2SMOTE',
                            'BPNN', 'BSMOTE', 'BWELM', 'CCR-ELM', 'ClusterSMOTE', 'CS-ELM', 'CSKELM',
                            'CSMOTE', 'DEBOHID', 'D-ENN', 'DNN', 'DT', 'EasyEnsemble', 'EFSVM', 'ELM',
                            'HABC-WELM', 'IWS-SMOTE', 'KELM', 'KMSMOTE', 'kNN', 'KNNOR', 'KWELM', 'LDA',
                            'MLP', 'MSOSS', 'MWMOTE', 'NB', 'NGSMOTE', 'Original', 'RF', 'ROS',
                            'RUSBoost', 'SGO', 'SMOTE', 'SMOTE-ENN', 'SMOTE-IPF', 'S-RSB',
                            'SSMOTE', 'S-TL', 'SVM', 'SyM', 'UBKELM-MV', 'UBKELM-SV', 'WELM', 'WKSMOTE', 
                            'XGB', 'BNB', 'UFIDSF', 'UCF', 'RENN', 'ENN', 'AKNN', 'GNB', 'PAU', 'OSS', 'TL', 
                            'RBU', 'RSDS', 'SDUS', 'UCF', 'MMTSSVM', 'FTSVM', 'RUS', 'TWSVM', 'RUTSVM',
                            'Tomek Links', 'Quadratic interpolation', 'Cubic interpolation', 'BENN',
                            'SVM-SMOTE', 'SMOTE-CDNN', 'RUS-SVM', 'Linear Classifiers', 'GRMM', 'ReliefFSVM',
                            'nonlinear-rbf-classifiers', 'TAE-GAN', 'CTGAN', 'SYMPROD', 'PSO', 'SVC', 'SMOTE-RST',
                            'GA', 'GWO', 'MVO', 'CDBH', 'AHC', 'EUSCHC', 'FRB+CHC', 'ICSPMH-GWO', 'DBSMOTE', 'DSMOTE',
                            'BGMM-SMOTE', 'Dirichlet ExtSMOTE', 'Gaussian SMOTE', 'Logistic Regression', 'ProWSyn',
                            'Distance ExtSMOTE', 'NN', 'RSMOTE', 'ABSMOTE', 'GBS', 'S+G+A', 'PCFS', 'Polynomial Curve Fitting',
                            'SMOTENC', 'GBO', 'SSG', 'SERP+', 'SL-D-Max', 'UFIDSF', 'UFIDSFeuclidean', 
                            'UFIDSFchebyshev', 'UFIDSFmanhattan', 'SMOTEBoost', 'BNSC', 'BRF', 'HSVM',
                            'ACFSVM', 'Gaussian kernel', 'RSVM', 'LSSVM', 'pinSVM', 'linear kernel', 'LrbssSVM',
                            'SMOTE-PSO', 'SMMO', 'SVM linear', 'SVM rbf', 'DB-MTD-SN', 'WIGNN', 'Graph Sage',
                            'CBWKELM', 'SV', 'MV', 'CBUS'}
        json = cargar_jsons_desde_directorio('JSON')
        analizar_jsons(json,tecnicas_validas)
    elif mode == '6':
        jsons = cargar_jsons_desde_directorio('JSON')
        techniques_dataset = get_techniques_xlsx(path='sheets/tecnicas.xlsx')
        datasets = get_datasets(path='sheets/Copia de datasets.xlsx')
        entries = {}
        for json in jsons:
            for dataset_id, dataset in json.get('datasets', {}).items():
                dataset_id = int(dataset_id)
                if dataset_id in datasets:
                    aux = dataset.get('auc', [])
                    values = aux.get('value', [])
                    techniques = aux.get('technique', [])
                    clasificadores = []
                    balanceadores = []
                    for index,_ in enumerate(techniques):
                        if techniques[index] is None or techniques[index] == "":
                            continue
                        types = re.split(r'[;,]', techniques[index])
                        for type_ in types:
                            if type_ in techniques_dataset:
                                if techniques_dataset[type_][1] == 1:
                                    clasificadores.append(type_)
                                elif techniques_dataset[type_][1] == 0:
                                    balanceadores.append(type_)
                        for balanceador in balanceadores:
                            if len(clasificadores) > 0:
                                for clasificador in clasificadores:
                                    entry = f"{' '.join(str(x) for x in datasets[dataset_id])} {techniques_dataset[clasificador][0]}"
                                    if entry not in entries:
                                        entries[entry] = (techniques_dataset[balanceador][0], values[index])
                                    elif entries[entry][1] < values[index]:
                                        entries[entry] = (techniques_dataset[balanceador][0], values[index])
                            else:
                                entry = f"{' '.join(str(x) for x in datasets[dataset_id])} 0"
                                if entry not in entries:
                                    entries[entry] = (techniques_dataset[balanceador][0], values[index])
                                elif entries[entry][1] < values[index]:
                                    entries[entry] = (techniques_dataset[balanceador][0], values[index])
        entries_list = list(entries.items())
        random.shuffle(entries_list)  # Mezcla aleatoriamente los datos

        # Calcular índice de corte para 2/3 entrenamiento
        corte = len(entries_list) * 2 // 3
        entrenamiento = entries_list[:corte]
        test = entries_list[corte:]

        # Encabezado común
        encabezado = ['Clase', 'Categóricos', 'Continuos', 'Discretos', 'Instancias', 'Desbalance', 'Clasificador', 'Balanceador']

        # Guardar conjunto de entrenamiento
        with open('train.csv', 'w', newline='') as archivo_train:
            writer = csv.writer(archivo_train)
            writer.writerow(encabezado)
            for clave, valor in entrenamiento:
                partes = [float(x) if '.' in x else int(x) for x in clave.split()]
                fila = partes + [valor[0]]
                writer.writerow(fila)

        # Guardar conjunto de test
        with open('test.csv', 'w', newline='') as archivo_test:
            writer = csv.writer(archivo_test)
            writer.writerow(encabezado)
            for clave, valor in test:
                partes = [float(x) if '.' in x else int(x) for x in clave.split()]
                fila = partes + [valor[0]]
                writer.writerow(fila)

